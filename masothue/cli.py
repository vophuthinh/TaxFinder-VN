# masothue/cli.py
# -*- coding: utf-8 -*-

"""
Command-line interface cho masothue package
"""

import argparse
import logging
import sys
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook, load_workbook

from masothue import MasothueClient, CompanySearchResult
from masothue.exceptions import (
    CaptchaRequiredError,
    NetworkError,
    ParseError,
    ValidationError,
    FileError
)
from masothue.utils import validate_excel_file, sanitize_filename
from masothue.formatters import format_company_details

logger = logging.getLogger(__name__)


def setup_logging(verbose: bool = False):
    """Cấu hình logging cho CLI"""
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )


def search_command(query: str, verbose: bool = False) -> int:
    """
    Tra cứu một công ty theo query (MST hoặc tên)
    
    Returns:
        0 nếu thành công, 1 nếu có lỗi
    """
    setup_logging(verbose)
    
    try:
        client = MasothueClient()
        results = client.search_companies(query)
        
        if not results:
            print(f"❌ Không tìm thấy công ty nào với query: {query}")
            return 1
        
        print(f"\n✅ Tìm thấy {len(results)} công ty:\n")
        for idx, result in enumerate(results, 1):
            print(f"{idx}. {result.name}")
            print(f"   MST: {result.tax_code}")
            if result.representative:
                print(f"   Người đại diện: {result.representative}")
            if result.tax_address:
                print(f"   Địa chỉ: {result.tax_address}")
            print()
        
        if len(results) == 1 and results[0].detail_url:
            print("📋 Lấy thông tin chi tiết...\n")
            try:
                details = client.get_company_details(results[0].detail_url)
                formatted = format_company_details(details)
                print(formatted)
            except CaptchaRequiredError:
                print("⚠️  Website yêu cầu CAPTCHA. Vui lòng tra cứu trên trình duyệt.")
                return 1
            except (NetworkError, ParseError) as e:
                print(f"❌ Lỗi khi lấy thông tin chi tiết: {e}")
                return 1
        
        return 0
        
    except CaptchaRequiredError:
        print("⚠️  Website yêu cầu CAPTCHA. Vui lòng tra cứu trên trình duyệt.")
        return 1
    except (NetworkError, ParseError) as e:
        print(f"❌ Lỗi: {e}")
        return 1
    except Exception as e:
        logger.exception("Unexpected error")
        print(f"❌ Lỗi không mong đợi: {e}")
        return 1


def batch_command(input_file: str, output_file: Optional[str] = None, verbose: bool = False) -> int:
    """
    Tra cứu hàng loạt từ file Excel
    
    Args:
        input_file: Đường dẫn file Excel đầu vào
        output_file: Đường dẫn file Excel đầu ra (nếu None, tự động tạo)
        verbose: Hiển thị log chi tiết
    
    Returns:
        0 nếu thành công, 1 nếu có lỗi
    """
    setup_logging(verbose)
    
    try:
        input_path = validate_excel_file(input_file)
        
        if not output_file:
            input_stem = input_path.stem
            output_file = str(input_path.parent / f"{input_stem}_results.xlsx")
        else:
            output_file = sanitize_filename(output_file)
            if not output_file.endswith('.xlsx'):
                output_file += '.xlsx'
        
        output_path = Path(output_file)
        
        print(f"📖 Đọc file: {input_path}")
        wb = load_workbook(input_path, data_only=True)
        ws = wb.active
        
        headers = []
        query_column_idx = None
        
        if ws.max_row > 0:
            for cell in ws[1]:
                headers.append(cell.value if cell.value else "")
        
        for idx, header in enumerate(headers):
            if header:
                header_lower = str(header).lower()
                if any(keyword in header_lower for keyword in ["tên", "name", "company", "mã số thuế", "mst", "tax"]):
                    query_column_idx = idx
                    query_column = header
                    break
        
        if query_column_idx is None:
            query_column_idx = 0
            query_column = headers[0] if headers else "Cột 1"
            print(f"⚠️  Không tìm thấy cột phù hợp. Sử dụng cột '{query_column}'")
        
        queries = []
        seen_queries = set()
        
        for row in ws.iter_rows(min_row=2, min_col=query_column_idx+1, max_col=query_column_idx+1, values_only=True):
            value = row[0]
            if value:
                value_str = str(value).strip()
                if value_str and value_str not in seen_queries:
                    if value_str.isdigit():
                        mst_length = len(value_str)
                        if mst_length in [10, 13] or (8 <= mst_length <= 15):
                            queries.append(value_str)
                            seen_queries.add(value_str)
                    else:
                        queries.append(value_str)
                        seen_queries.add(value_str)
        
        if not queries:
            print("❌ Không tìm thấy dữ liệu để tra cứu trong file Excel.")
            return 1
        
        print(f"📊 Tìm thấy {len(queries)} công ty để tra cứu\n")
        
        client = MasothueClient()
        results = []
        errors = []
        
        for idx, query in enumerate(queries, 1):
            print(f"[{idx}/{len(queries)}] Tra cứu: {query}...", end=" ", flush=True)
            
            try:
                search_results = client.search_companies(query)
                
                if search_results:
                    result = search_results[0]
                    if result.detail_url:
                        try:
                            details = client.get_company_details(result.detail_url)
                            results.append({
                                'query': query,
                                'result': result,
                                'details': details
                            })
                            print("✅")
                        except CaptchaRequiredError:
                            print("⚠️  CAPTCHA")
                            errors.append({'query': query, 'error': 'CAPTCHA required'})
                        except (NetworkError, ParseError) as e:
                            print(f"❌ {e}")
                            errors.append({'query': query, 'error': str(e)})
                    else:
                        results.append({
                            'query': query,
                            'result': result,
                            'details': None
                        })
                        print("✅")
                else:
                    print("❌ Không tìm thấy")
                    errors.append({'query': query, 'error': 'Not found'})
                    
            except CaptchaRequiredError:
                print("⚠️  CAPTCHA")
                errors.append({'query': query, 'error': 'CAPTCHA required'})
            except (NetworkError, ParseError) as e:
                print(f"❌ {e}")
                errors.append({'query': query, 'error': str(e)})
            except Exception as e:
                logger.exception(f"Error processing {query}")
                print(f"❌ {e}")
                errors.append({'query': query, 'error': str(e)})
        
        print(f"\n💾 Ghi kết quả vào: {output_path}")
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = "Kết quả tra cứu"
        
        from masothue.formatters import make_result_row
        
        sample_row = make_result_row("sample")
        headers_out = list(sample_row.keys())
        ws_out.append(headers_out)
        
        for item in results:
            result = item['result']
            details = item.get('details')
            row_dict = make_result_row(item['query'], result=result, details=details)
            row = [row_dict.get(key, '') for key in headers_out]
            ws_out.append(row)
        
        for error_item in errors:
            row_dict = make_result_row(error_item['query'], error=error_item['error'])
            row = [row_dict.get(key, '') for key in headers_out]
            ws_out.append(row)
        
        wb_out.save(output_path)
        
        print(f"\n✅ Hoàn thành!")
        print(f"   - Thành công: {len(results)}")
        print(f"   - Lỗi: {len(errors)}")
        print(f"   - File kết quả: {output_path}")
        
        return 0
        
    except FileError as e:
        print(f"❌ Lỗi file: {e}")
        return 1
    except ValidationError as e:
        print(f"❌ Lỗi validation: {e}")
        return 1
    except Exception as e:
        logger.exception("Unexpected error")
        print(f"❌ Lỗi không mong đợi: {e}")
        return 1


def main():
    """Entry point cho CLI"""
    parser = argparse.ArgumentParser(
        description="Tra cứu mã số thuế công ty từ masothue.com",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ví dụ:
  %(prog)s search --query "1234567890"
  %(prog)s search --query "Công ty ABC"
  %(prog)s batch input.xlsx output.xlsx
  %(prog)s batch input.xlsx  # Tự động tạo output file
        """
    )
    
    parser.add_argument(
        '-v', '--verbose',
        action='store_true',
        help='Hiển thị log chi tiết'
    )
    
    subparsers = parser.add_subparsers(dest='command', help='Lệnh cần thực hiện')
    
    search_parser = subparsers.add_parser('search', help='Tra cứu một công ty')
    search_parser.add_argument(
        '--query', '-q',
        required=True,
        help='Mã số thuế hoặc tên công ty cần tra cứu'
    )
    
    batch_parser = subparsers.add_parser('batch', help='Tra cứu hàng loạt từ file Excel')
    batch_parser.add_argument(
        'input_file',
        help='Đường dẫn file Excel đầu vào'
    )
    batch_parser.add_argument(
        'output_file',
        nargs='?',
        default=None,
        help='Đường dẫn file Excel đầu ra (tùy chọn, mặc định: input_results.xlsx)'
    )
    
    args = parser.parse_args()
    
    if not args.command:
        parser.print_help()
        return 1
    
    if args.command == 'search':
        return search_command(args.query, args.verbose)
    elif args.command == 'batch':
        return batch_command(args.input_file, args.output_file, args.verbose)
    else:
        parser.print_help()
        return 1


if __name__ == "__main__":
    sys.exit(main())

