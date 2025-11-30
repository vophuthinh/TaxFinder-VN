# excel_service.py
# -*- coding: utf-8 -*-

"""
Service layer cho các thao tác Excel.
Tách biệt logic xử lý file Excel khỏi UI để tuân thủ Single Responsibility Principle.
"""

import logging
from pathlib import Path
from datetime import datetime
from typing import List, Optional, Tuple, Dict, Any

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

from masothue.utils import clean_tax_code, is_valid_tax_code, validate_excel_file
from masothue.exceptions import FileError, ValidationError

logger = logging.getLogger(__name__)


def read_queries_from_excel(
    file_path: str,
    query_column_idx: Optional[int] = None
) -> Tuple[List[str], Optional[int], Optional[str]]:
    """
    Đọc danh sách queries (MST hoặc tên công ty) từ file Excel.
    
    Args:
        file_path: Đường dẫn đến file Excel
        query_column_idx: Chỉ số cột cần đọc (0-based). Nếu None, sẽ tự động phát hiện.
        
    Returns:
        Tuple (queries, column_idx, column_name):
            - queries: Danh sách queries đã được clean và validate
            - column_idx: Chỉ số cột đã sử dụng (0-based)
            - column_name: Tên cột đã sử dụng
            
    Raises:
        FileError: Nếu không đọc được file
        ValidationError: Nếu file không hợp lệ
    """
    try:
        validated_path = validate_excel_file(file_path)
        file_path = str(validated_path)
    except ValidationError as e:
        raise ValidationError(f"File Excel không hợp lệ: {e.message}")
    except FileError as e:
        raise FileError(f"Không thể đọc file Excel: {e.message}")
    
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
    except Exception as e:
        raise FileError(f"Lỗi khi mở file Excel: {str(e)}")
    
    headers: List[str] = []
    if ws.max_row > 0:
        for cell in ws[1]:
            headers.append(cell.value if cell.value else "")
    
    column_name: Optional[str] = None
    if query_column_idx is None:
        query_column_idx, column_name = detect_query_column(headers)
        if query_column_idx is None:
            raise ValidationError("Không tìm thấy cột chứa MST hoặc tên công ty. Vui lòng chọn cột thủ công.")
    else:
        if query_column_idx < 0 or query_column_idx >= len(headers):
            raise ValidationError(f"Chỉ số cột không hợp lệ: {query_column_idx}")
        column_name = headers[query_column_idx] if query_column_idx < len(headers) else None
    
    queries: List[str] = []
    seen_queries = set()
    
    try:
        for row in ws.iter_rows(min_row=2, min_col=query_column_idx+1, max_col=query_column_idx+1, values_only=True):
            value = row[0]
            if value:
                value_str = str(value).strip()
                if not value_str or value_str in seen_queries:
                    continue
                
                cleaned_mst = clean_tax_code(value_str)
                
                if cleaned_mst and is_valid_tax_code(cleaned_mst):
                    queries.append(cleaned_mst)
                    seen_queries.add(cleaned_mst)
                    if len(cleaned_mst) not in [10, 13]:
                        logger.warning(
                            f"MST có độ dài không chuẩn ({len(cleaned_mst)} chữ số): "
                            f"{cleaned_mst} (từ '{value_str}') - vẫn sẽ tra cứu"
                        )
                elif cleaned_mst:
                    logger.warning(
                        f"MST không hợp lệ ({len(cleaned_mst)} chữ số): "
                        f"{cleaned_mst} (từ '{value_str}') - bỏ qua"
                    )
                else:
                    sanitized = value_str.strip()
                    if sanitized:
                        queries.append(sanitized)
                        seen_queries.add(sanitized)
    except Exception as e:
        raise FileError(f"Lỗi khi đọc dữ liệu từ Excel: {str(e)}")
    
    return queries, query_column_idx, column_name


def detect_query_column(headers: List[str]) -> Tuple[Optional[int], Optional[str]]:
    """
    Tự động phát hiện cột chứa MST hoặc tên công ty từ headers.
    
    Args:
        headers: Danh sách tên cột từ hàng đầu tiên của Excel
        
    Returns:
        Tuple (column_idx, column_name):
            - column_idx: Chỉ số cột (0-based) hoặc None nếu không tìm thấy
            - column_name: Tên cột hoặc None nếu không tìm thấy
    """
    # Keywords chính xác cho MST (ưu tiên cao nhất)
    mst_keywords_exact = ["mã số thuế", "mst", "tax code", "tax_code", "mã số thuế doanh nghiệp"]
    # Keywords chính xác cho tên công ty (ưu tiên cao)
    company_keywords_exact = ["tên công ty", "tên doanh nghiệp", "company name", "company_name", "tên"]
    # Keywords phụ (ưu tiên thấp hơn, cần kiểm tra kỹ)
    company_keywords_loose = ["name", "company"]
    
    query_column_idx: Optional[int] = None
    query_column: Optional[str] = None
    
    # Tìm với keywords chính xác
    for idx, header in enumerate(headers):
        if not header:
            continue
        
        header_str = str(header).strip()
        header_lower = header_str.lower()
        
        for keyword in mst_keywords_exact:
            if keyword == header_lower or keyword in header_lower:
                if "người nhận" not in header_lower and "khách hàng" not in header_lower:
                    query_column_idx = idx
                    query_column = header_str
                    return query_column_idx, query_column
        
        for keyword in company_keywords_exact:
            if keyword == header_lower or keyword in header_lower:
                if all(exclude not in header_lower for exclude in ["người nhận", "khách hàng", "người liên hệ", "người gửi"]):
                    query_column_idx = idx
                    query_column = header_str
                    return query_column_idx, query_column
    
    for idx, header in enumerate(headers):
        if not header:
            continue
        
        header_str = str(header).strip()
        header_lower = header_str.lower()
        
        for keyword in company_keywords_loose:
            if keyword in header_lower:
                if all(exclude not in header_lower for exclude in [
                    "người nhận", "khách hàng", "người liên hệ", "người gửi",
                    "recipient", "customer", "contact"
                ]):
                    query_column_idx = idx
                    query_column = header_str
                    return query_column_idx, query_column
    
    return None, None


def export_results_to_excel(
    data: List[Dict[str, Any]],
    file_path: str,
    metadata: Optional[Dict[str, Any]] = None
) -> None:
    """
    Xuất kết quả tra cứu ra file Excel.
    
    Args:
        data: Danh sách dictionaries chứa kết quả tra cứu
        file_path: Đường dẫn file Excel để lưu
        metadata: Dictionary chứa metadata (timestamp, source, count, etc.)
        
    Raises:
        FileError: Nếu không ghi được file
        
    Performance Note:
        - Với file nhỏ (< 5000 dòng): Cách hiện tại (ws.cell()) là ổn
        - Với file lớn (> 50.000 dòng): Nên cân nhắc:
          * Dùng pandas: df.to_excel() (nhanh hơn nhiều)
          * Hoặc openpyxl write_only=True mode (tiết kiệm memory)
          * Hoặc batch write (append nhiều rows cùng lúc)
    """
    if not data:
        raise ValueError("Không có dữ liệu để xuất")
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Kết quả tra cứu"
        
        row = 1
        if metadata:
            if "timestamp" in metadata:
                ws.cell(row=row, column=1, value="Thời gian xuất:")
                ws.cell(row=row, column=2, value=metadata["timestamp"])
                row += 1
            if "source" in metadata:
                ws.cell(row=row, column=1, value="Nguồn tra cứu:")
                ws.cell(row=row, column=2, value=metadata["source"])
                row += 1
            if "count" in metadata:
                ws.cell(row=row, column=1, value="Số lượng kết quả:")
                ws.cell(row=row, column=2, value=metadata["count"])
                row += 1
        
        header_row = row + 1
        if data:
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=header_row, column=col_idx, value=header)
                ws.column_dimensions[get_column_letter(col_idx)].width = 20
            
            data_start_row = header_row + 1
            for row_idx, row_data in enumerate(data, data_start_row):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(file_path)
        logger.info(f"Đã xuất {len(data)} kết quả ra file Excel: {file_path}")
        
    except Exception as e:
        raise FileError(f"Lỗi khi ghi file Excel: {str(e)}")

