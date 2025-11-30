# masothue_app.py
# -*- coding: utf-8 -*-

import logging
import os
import re
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import webbrowser
import threading
import queue
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any, TypedDict, Union

from openpyxl import load_workbook

from masothue import MasothueClient, CompanySearchResult
from masothue.utils import is_valid_tax_code, sanitize_query, sanitize_filename
from masothue.excel_service import read_queries_from_excel, export_results_to_excel
from masothue.batch_worker import BatchWorker
from masothue.exceptions import (
    CaptchaRequiredError,
    NetworkError,
    ParseError,
    ValidationError,
    FileError,
    CancelledError
)
from masothue.config import DEFAULT_RATE_LIMIT, CACHE_ENABLED, CACHE_EXPIRY_DAYS
from masothue.theme import Theme
from masothue.constants import (
    MSG_SEARCHING,
    MSG_BATCH_COMPLETE,
    MSG_BATCH_CANCELLED,
    MSG_BATCH_STOPPED,
    ERR_FILE_READ,
    MSG_EMPTY_STATE,
    MSG_NO_LOG_FILE,
    MSG_NO_EXCEL_DATA,
    MSG_NO_SEARCH_RESULTS,
    MSG_SEARCH_SUCCESS,
    MSG_SEARCH_ERROR,
    MSG_NETWORK_ERROR,
    MSG_NETWORK_ERROR_DETAIL,
    MSG_VALIDATION_ERROR,
    MSG_CAPTCHA_REQUIRED,
    MSG_CAPTCHA_REQUIRED_DETAIL,
    MSG_SEARCHING_MST,
    MSG_NO_DETAIL_INFO,
    MSG_CLEAR_RESULTS,
    MSG_CANNOT_OPEN_LOG,
    MSG_INVALID_SETTINGS,
    MSG_INVALID_NUMBER,
    MSG_EXPORT_ERROR,
    MSG_SEARCH_ERROR_DETAIL,
    MSG_FILE_INVALID,
    MSG_FILE_READ_ERROR,
    MSG_COLUMN_SELECTION,
    MSG_COLUMN_SELECTION_INFO,
    MSG_NO_COLUMNS,
    MSG_NO_COLUMNS_DETAIL,
    MSG_COLUMN_NOT_SELECTED,
    MSG_COLUMN_NOT_SELECTED_DETAIL,
    CONFIRM_EXPORT,
    CONFIRM_EXIT,
    SUCCESS_BATCH_COMPLETE,
    TITLE_INFO,
    TITLE_ERROR,
    TITLE_WARNING,
    TITLE_CONFIRM,
    TITLE_CONFIRM_CANCEL,
    TITLE_CAPTCHA,
    TITLE_NETWORK_ERROR,
    TITLE_VALIDATION_ERROR,
    TITLE_FILE_ERROR,
    TITLE_EXPORT_ERROR,
    TITLE_EXIT,
    ERR_FILE_WRITE,
    ERR_CAPTCHA,
    ERR_INVALID_INPUT,
    ERR_INVALID_FILE_PATH,
    CONFIRM_EXIT_BATCH,
    CONFIRM_CANCEL_BATCH,
    SUCCESS_EXPORT,
    SUCCESS_BATCH_CANCELLED,
    ALLOWED_EXCEL_EXTENSIONS,
    MAX_FILE_SIZE_MB,
    MAX_QUERY_LENGTH
)

logger = logging.getLogger(__name__)


class CompanyDetailRecord(TypedDict, total=False):
    """
    Cấu trúc dữ liệu chi tiết công ty lưu trong item_data_map.
    Tất cả fields đều là Optional[str] (có thể là empty string).
    
    Lưu ý về Type Safety:
        - total=False nghĩa là tất cả fields đều optional
        - Khi access data['name'], nếu key không tồn tại sẽ gây KeyError
        - Code hiện tại đã an toàn vì:
          * Luôn dùng .get() với default value: data.get('name', '')
          * Hoặc khởi tạo dict với tất cả keys có default value
        - Nếu refactor sau này, nhớ luôn dùng .get() hoặc check 'key' in data
    """
    name: str
    tax_code: str
    representative: str
    tax_address: str
    address: str
    phone: str
    status: str
    operation_date: str
    managed_by: str
    business_type: str
    main_business: str
    other_businesses: Union[str, List[str]]


class CompanyDetails(TypedDict, total=False):
    """
    Cấu trúc dữ liệu từ get_company_details() trong masothue/client.py.
    Tất cả fields đều là Optional.
    """
    name: Optional[str]
    tax_code: Optional[str]
    representative: Optional[str]
    tax_address: Optional[str]
    address: Optional[str]
    phone: Optional[str]
    status: Optional[str]
    operation_date: Optional[str]
    managed_by: Optional[str]
    business_type: Optional[str]
    main_business: Optional[str]
    other_businesses: Union[Optional[str], Optional[List[str]]]


BatchResultDict = Dict[str, str]
CompanyDetailDict = CompanyDetails


class MasothueApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()

        self.title("Tra cứu mã số thuế")
        
        # Thử set icon nếu có
        try:
            icon_path = Path(__file__).parent / "masothue.ico"
            if icon_path.exists():
                self.iconbitmap(str(icon_path))
        except (OSError, tk.TclError):
            pass
        
        self.progress_queue = queue.Queue()
        
        self.update_idletasks()
        width = 1200
        height = 750
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")
        
        self.last_dir = None

        self.client = MasothueClient(
            max_requests=DEFAULT_RATE_LIMIT["max_requests"],
            time_window=DEFAULT_RATE_LIMIT["time_window"],
            min_delay=DEFAULT_RATE_LIMIT["min_delay"]
        )

        self._setup_style()
        self._build_ui()
        
        self.protocol("WM_DELETE_WINDOW", self._on_closing)
        
        self._poll_progress_queue()
    
    def _setup_style(self):
        """Thiết lập theme và font đồng bộ với màu sắc đẹp"""
        style = ttk.Style(self)
        self._setup_base_style(style)
        self._setup_treeview_style(style)
        self._setup_button_styles(style)
        self._setup_entry_style(style)
        self._setup_progressbar_style(style)
        
    def _setup_base_style(self, style: ttk.Style):
        """Thiết lập base style: theme, font, background"""
        try:
            style.theme_use("clam")
        except:
            style.theme_use("default")
        
        default_font = ("Segoe UI", 10)
        self.option_add("*TLabel.Font", default_font)
        self.option_add("*TButton.Font", default_font)
        self.option_add("*TEntry.Font", default_font)
        self.option_add("*TFrame.Font", default_font)
        
        self.configure(bg=Theme.BG_MAIN)
        
        style.configure("TLabelframe",
                       background=Theme.BG_WHITE,
                       borderwidth=2,
                       relief="solid",
                       bordercolor=Theme.BORDER_LIGHT)
        style.configure("TLabelframe.Label",
                       font=("Segoe UI", 12, "bold"),
                       background=Theme.BG_WHITE,
                       foreground=Theme.TEXT_PRIMARY)
        
        style.configure("Search.TLabel",
                       background=Theme.BG_WHITE,
                       foreground=Theme.TEXT_PRIMARY,
                       font=("Segoe UI", 10))
        style.configure("Status.TLabel",
                       background=Theme.BG_WHITE,
                       foreground=Theme.PRIMARY,
                       font=("Segoe UI", 9, "italic"))
    
    def _setup_treeview_style(self, style: ttk.Style) -> None:
        """Thiết lập style cho Treeview"""
        style.configure("Treeview.Heading", 
                       font=("Segoe UI", 10, "bold"),
                       background=Theme.BG_HEADER,
                       foreground=Theme.TEXT_WHITE,
                       relief="flat",
                       borderwidth=0,
                       padding=(8, 8))
        style.map("Treeview.Heading",
                 background=[("active", "#34495e"),
                            ("pressed", Theme.PRIMARY_DARKER)])
        
        style.configure("Treeview",
                       rowheight=28,
                       background=Theme.BG_WHITE,
                       foreground=Theme.TEXT_PRIMARY,
                       fieldbackground=Theme.BG_WHITE,
                       borderwidth=0,
                       font=("Segoe UI", 9))
        
        style.map("Treeview",
                 background=[("selected", Theme.SELECTED)],
                 foreground=[("selected", Theme.SELECTED_TEXT)])
        
    def _setup_button_styles(self, style: ttk.Style) -> None:
        """Thiết lập các style cho Button"""
        style.configure("Custom.TButton",
                       font=("Segoe UI", 10, "bold"),
                       padding=(18, 10),
                       relief="flat",
                       borderwidth=0)
        style.map("Custom.TButton",
                 background=[("active", "#2980b9"),
                            ("!active", "#3498db")],
                 foreground=[("active", "white"),
                           ("!active", "white")])
        
        style.configure("Primary.TButton",
                       font=("Segoe UI", 11, "bold"),
                       padding=(25, 12),
                       relief="flat",
                       borderwidth=0)
        style.map("Primary.TButton",
                 background=[("active", "#27ae60"),
                            ("!active", "#2ecc71")],
                 foreground=[("active", "white"),
                           ("!active", "white")])
        
        style.configure("Secondary.TButton",
                       font=("Segoe UI", 9),
                       padding=(15, 8),
                       relief="flat",
                       borderwidth=0)
        style.map("Secondary.TButton",
                 background=[("active", "#95a5a6"),
                            ("!active", "#bdc3c7")],
                 foreground=[("active", "white"),
                           ("!active", "#2c3e50")])
        
        style.configure("Cancel.TButton",
                       font=("Segoe UI", 10, "bold"),
                       padding=(20, 10),
                       relief="flat",
                       borderwidth=0)
        style.map("Cancel.TButton",
                 background=[("active", "#c0392b"),
                            ("!active", "#e74c3c")],
                 foreground=[("active", "white"),
                           ("!active", "white")])
    
    def _setup_entry_style(self, style: ttk.Style) -> None:
        """Thiết lập style cho Entry"""
        style.configure("Custom.TEntry",
                       fieldbackground="#ffffff",
                       borderwidth=2,
                       relief="solid",
                       bordercolor="#bdc3c7",
                       padding=8,
                       font=("Segoe UI", 10))
        style.map("Custom.TEntry",
                 bordercolor=[("focus", "#3498db"),
                             ("!focus", "#bdc3c7")])
        
    def _setup_progressbar_style(self, style: ttk.Style) -> None:
        """Thiết lập style cho Progressbar"""
        style.configure("TProgressbar",
                       background="#3498db",
                       troughcolor="#ecf0f1",
                       borderwidth=0,
                       lightcolor="#3498db",
                       darkcolor="#2980b9",
                       thickness=25)
        
        style.layout("ETL.TProgressbar", [
            ("Horizontal.Progressbar.trough", {
                "sticky": "nswe",
                "children": [
                    ("Horizontal.Progressbar.pbar", {
                        "side": "left",
                        "sticky": "we"
                    })
                ]
            })
        ])
        style.configure("ETL.TProgressbar",
                       background="#3498db",
                       troughcolor="#ecf0f1",
                       borderwidth=0,
                       lightcolor="#3498db",
                       darkcolor="#2980b9",
                       thickness=8)

    def _build_ui(self) -> None:
        """Xây dựng toàn bộ UI"""
        self._build_menu()
        self._build_header()
        self._build_quick_search_frame()
        self._build_batch_frame()
        self._build_results_frame()
        self._build_detail_frame()
        
        self._batch_state: str = "IDLE"
        self._batch_cancelled: bool = False
        self._is_batch_running: bool = False
        self.batch_results: List[BatchResultDict] = []
    
    def _build_menu(self) -> None:
        """Xây dựng menu bar"""
        menubar = tk.Menu(self)
        self.config(menu=menubar)

        # Menu Cài đặt / Tác vụ chính
        settings_menu = tk.Menu(menubar, tearoff=0)
        settings_menu.add_command(label="Xóa kết quả", command=self.on_clear_results)
        settings_menu.add_command(label="Làm mới", command=self.on_refresh)
        settings_menu.add_separator()
        settings_menu.add_command(label="Cài đặt nâng cao...", command=self._show_advanced_settings)
        menubar.add_cascade(label="Cài đặt", menu=settings_menu)
        
        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="Mở log", command=self._open_log_file)
        help_menu.add_separator()
        help_menu.add_command(label="Về ứng dụng", command=self._show_about)
        menubar.add_cascade(label="Trợ giúp", menu=help_menu)
    
    def _build_header(self) -> None:
        """Xây dựng header bar trên cùng"""
        header = tk.Frame(self, bg="#2c3e50")
        header.pack(side="top", fill="x")
        
        title_container = tk.Frame(header, bg=Theme.BG_HEADER)
        title_container.pack(side="left", padx=20, pady=(8, 8))
        
        title = tk.Label(
            title_container,
            text="Tra cứu mã số thuế doanh nghiệp",
            font=("Segoe UI", 16, "bold"),
            bg=Theme.BG_HEADER,
            fg=Theme.TEXT_HEADER,
            anchor="w"
        )
        title.pack(side="top", anchor="w")
        
        subtitle = tk.Label(
            title_container,
            text="Tìm nhanh – Lưu cache – Xuất Excel",
            font=("Segoe UI", 9),
            bg=Theme.BG_HEADER,
            fg=Theme.TEXT_HEADER_SUBTITLE,
            anchor="w"
        )
        subtitle.pack(side="top", anchor="w", pady=(2, 0))
    
    def on_clear_results(self) -> None:
        """Xóa kết quả tra cứu"""
        if not hasattr(self, 'tree'):
            return  # Chưa khởi tạo UI
        
        self.tree.delete(*self.tree.get_children())
        self.item_url_map.clear()
        self.item_data_map.clear()
        self.batch_results = []
        
        # Reset panel chi tiết
        if hasattr(self, 'detail_labels'):
            for label in self.detail_labels.values():
                label.config(text="", fg=Theme.TEXT_PRIMARY)
        
        self.set_status("✓ Đã xóa kết quả", force_update=True)
        
        # Hiển thị empty state
        if hasattr(self, 'empty_label'):
            self.empty_label.config(
                text=MSG_EMPTY_STATE
            )
            self.empty_label.place(relx=0.5, rely=0.5, anchor="center")
    
    def on_refresh(self) -> None:
        """Làm mới ứng dụng"""
        self.on_clear_results()
        self.set_status("✓ Sẵn sàng", force_update=True)
    
    def _open_log_file(self) -> None:
        """Mở file log trong trình soạn thảo mặc định"""
        
        log_dir = Path("logs")
        if not log_dir.exists():
            messagebox.showinfo(TITLE_INFO, MSG_NO_LOG_FILE)
            return
        
        log_files = sorted(log_dir.glob("masothue_*.log"), key=os.path.getmtime, reverse=True)
        if not log_files:
            messagebox.showinfo(TITLE_INFO, MSG_NO_LOG_FILE)
            return
        
        log_file = log_files[0]
        try:
            os.startfile(str(log_file))
        except (OSError, FileNotFoundError, PermissionError) as e:
            messagebox.showerror(TITLE_ERROR, MSG_CANNOT_OPEN_LOG.format(error=e))
    
    def _show_advanced_settings(self) -> None:
        """Hiển thị dialog cài đặt nâng cao"""
        
        dialog = tk.Toplevel(self)
        dialog.title("Cài đặt nâng cao")
        dialog.geometry("500x550")
        dialog.transient(self)
        dialog.grab_set()
        
        current_max_requests = self.client.rate_limiter.max_requests
        current_time_window = self.client.rate_limiter.time_window
        current_min_delay = self.client.rate_limiter.min_delay
        current_cache_enabled = self.client.cache_enabled
        current_cache_expiry = self.client.file_cache.expiry_days if self.client.file_cache else CACHE_EXPIRY_DAYS
        
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill="both", expand=True)
        
        rate_frame = ttk.LabelFrame(main_frame, text="Giới hạn tần suất (Rate Limiting)", padding=10)
        rate_frame.pack(fill="x", pady=(0, 15))
        
        ttk.Label(rate_frame, text="Số request tối đa:").grid(row=0, column=0, sticky="w", pady=5)
        max_requests_var = tk.StringVar(value=str(current_max_requests))
        max_requests_entry = ttk.Entry(rate_frame, textvariable=max_requests_var, width=15)
        max_requests_entry.grid(row=0, column=1, sticky="w", padx=10, pady=5)
        ttk.Label(rate_frame, text="(mặc định: 10)", foreground="gray").grid(row=0, column=2, sticky="w")
        
        ttk.Label(rate_frame, text="Cửa sổ thời gian (giây):").grid(row=1, column=0, sticky="w", pady=5)
        time_window_var = tk.StringVar(value=str(current_time_window))
        time_window_entry = ttk.Entry(rate_frame, textvariable=time_window_var, width=15)
        time_window_entry.grid(row=1, column=1, sticky="w", padx=10, pady=5)
        ttk.Label(rate_frame, text="(mặc định: 60)", foreground="gray").grid(row=1, column=2, sticky="w")
        
        ttk.Label(rate_frame, text="Độ trễ tối thiểu (giây):").grid(row=2, column=0, sticky="w", pady=5)
        min_delay_var = tk.StringVar(value=str(current_min_delay))
        min_delay_entry = ttk.Entry(rate_frame, textvariable=min_delay_var, width=15)
        min_delay_entry.grid(row=2, column=1, sticky="w", padx=10, pady=5)
        ttk.Label(rate_frame, text="(mặc định: 1.0)", foreground="gray").grid(row=2, column=2, sticky="w")
        
        cache_frame = ttk.LabelFrame(main_frame, text="Cache", padding=10)
        cache_frame.pack(fill="x", pady=(0, 15))
        
        cache_enabled_var = tk.BooleanVar(value=current_cache_enabled)
        cache_check = ttk.Checkbutton(cache_frame, text="Bật cache", variable=cache_enabled_var)
        cache_check.pack(anchor="w", pady=5)
        
        ttk.Label(cache_frame, text="Thời gian hết hạn cache (ngày):").pack(anchor="w", pady=(10, 5))
        cache_expiry_var = tk.StringVar(value=str(current_cache_expiry))
        cache_expiry_entry = ttk.Entry(cache_frame, textvariable=cache_expiry_var, width=15)
        cache_expiry_entry.pack(anchor="w", pady=5)
        ttk.Label(cache_frame, text="(mặc định: 7)", foreground="gray").pack(anchor="w")
        
        hint_label = tk.Label(
            main_frame,
            text="💡 Lưu ý: Thay đổi sẽ áp dụng cho các lần tra cứu tiếp theo.\n"
                 "Giảm tần suất request và tăng delay giúp giảm khả năng bị CAPTCHA.",
            foreground=Theme.TEXT_SECONDARY,
            font=("Segoe UI", 9, "italic"),
            justify="left",
            wraplength=450
        )
        hint_label.pack(fill="x", pady=(0, 15))
        
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x")
        
        def apply_settings():
            try:
                max_requests = int(max_requests_var.get())
                time_window = int(time_window_var.get())
                min_delay = float(min_delay_var.get())
                cache_enabled = cache_enabled_var.get()
                cache_expiry = int(cache_expiry_var.get())
                
                if max_requests < 1 or time_window < 1 or min_delay < 0 or cache_expiry < 1:
                    messagebox.showerror("Lỗi", "Các giá trị phải là số dương.")
                    return
                
                self.client = MasothueClient(
                    max_requests=max_requests,
                    time_window=time_window,
                    min_delay=min_delay,
                    enable_cache=cache_enabled,
                    cache_expiry_days=cache_expiry
                )
                
                messagebox.showinfo("Thành công", "Đã cập nhật cài đặt thành công!")
                dialog.destroy()
            except ValueError:
                messagebox.showerror(TITLE_ERROR, MSG_INVALID_NUMBER)
        
        def reset_to_default():
            max_requests_var.set(str(DEFAULT_RATE_LIMIT["max_requests"]))
            time_window_var.set(str(DEFAULT_RATE_LIMIT["time_window"]))
            min_delay_var.set(str(DEFAULT_RATE_LIMIT["min_delay"]))
            cache_enabled_var.set(CACHE_ENABLED)
            cache_expiry_var.set(str(CACHE_EXPIRY_DAYS))
        
        ttk.Button(button_frame, text="Đặt lại mặc định", command=reset_to_default).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Hủy", command=dialog.destroy).pack(side="right", padx=5)
        ttk.Button(button_frame, text="Áp dụng", command=apply_settings).pack(side="right", padx=5)
    
    def _show_about(self) -> None:
        """Hiển thị thông tin về ứng dụng"""
        messagebox.showinfo(
            "Về ứng dụng",
            "Tra cứu mã số thuế\n\n"
            "Phiên bản: 1.1\n\n"
            "Ứng dụng tra cứu thông tin công ty từ masothue.com\n"
            "Hỗ trợ tra cứu nhanh và tra cứu hàng loạt từ file Excel."
        )
    
    def _build_quick_search_frame(self) -> None:
        """Xây dựng frame tra cứu nhanh"""
        quick_frame_container = tk.Frame(self, bg="#f0f2f5")
        quick_frame_container.pack(side="top", fill="x", padx=15, pady=(15, 10))
        
        quick_frame = ttk.LabelFrame(quick_frame_container, text="⚡ Tra cứu nhanh", padding=20)
        quick_frame.pack(fill="both", expand=True)

        # Dùng grid layout để tránh chồng chữ trên màn hình hẹp
        # Cấu hình cột để entry co giãn
        quick_frame.columnconfigure(1, weight=1)

        # Hàng 1: Label | Entry (co giãn) | Button
        # Dùng ttk.Label thay vì tk.Label để đồng bộ style
        lbl = ttk.Label(quick_frame, text="Nhập tên công ty hoặc mã số thuế:", 
                       style="Search.TLabel")
        lbl.grid(row=0, column=0, sticky="w", padx=(0, 8))

        self.query_var = tk.StringVar()
        self.query_entry = ttk.Entry(quick_frame, textvariable=self.query_var, style="Custom.TEntry")
        self.query_entry.grid(row=0, column=1, sticky="ew", padx=8)
        self.query_entry.bind("<Return>", lambda e: self.on_search())
        self.query_entry.bind("<KeyRelease>", self._on_query_change)

        self.search_button = ttk.Button(quick_frame, text="⚡ Tra cứu", command=self.on_search, style="Primary.TButton")
        self.search_button.grid(row=0, column=2, padx=(8, 0))

        # Hàng 2: Status label (colspan 3, căn trái)
        # Dùng ttk.Label với style Status.TLabel để đồng bộ
        self.query_status_label = ttk.Label(
            quick_frame,
            text="",
            style="Status.TLabel"
        )
        self.query_status_label.grid(row=1, column=0, columnspan=3, sticky="w", pady=(8, 0))
        
    def _build_batch_frame(self) -> None:
        """Xây dựng frame tra cứu hàng loạt"""
        batch_frame_container = tk.Frame(self, bg=Theme.BG_MAIN)
        batch_frame_container.pack(side="top", fill="x", padx=15, pady=(0, 15))
        
        batch_frame = ttk.LabelFrame(batch_frame_container, text="📈 Tra cứu hàng loạt (Excel)", padding=20)
        batch_frame.pack(fill="both", expand=True)
        
        self.import_button = ttk.Button(batch_frame, text="📂 Nhập Excel", command=self.on_import_excel, style="Custom.TButton")
        self.import_button.pack(side="left", padx=5)
        
        self.export_button = ttk.Button(batch_frame, text="💾 Xuất Excel", command=self.on_export_excel, style="Custom.TButton")
        self.export_button.pack(side="left", padx=5)
        
        hint_excel = tk.Label(
            batch_frame, 
            text="💡 Lưu ý: File Excel cần có cột chứa mã số thuế hoặc tên công ty",
            foreground=Theme.TEXT_SECONDARY,
            font=("Segoe UI", 9, "italic"),
            background=Theme.BG_WHITE
        )
        hint_excel.pack(side="left", padx=(25, 0))

    def _build_results_frame(self) -> None:
        """Xây dựng frame kết quả tra cứu (treeview)"""
        # Chia phần giữa thành 2 cột với PanedWindow
        self.main_pane = ttk.PanedWindow(self, orient="horizontal")
        self.main_pane.pack(fill="both", expand=True, padx=10, pady=5)

        left_frame = ttk.Frame(self.main_pane)
        self.main_pane.add(left_frame, weight=3)

        # Label hiển thị số kết quả
        results_header = tk.Frame(left_frame, bg=Theme.BG_RESULTS_HEADER, relief="flat")
        results_header.pack(side="top", fill="x", padx=5, pady=(0, 10))
        # KHÔNG pack_propagate(False) nữa
        
        # Label cố định "Kết quả tra cứu" - không cập nhật trạng thái
        results_label = tk.Label(
            results_header, 
            text="💼 Kết quả tra cứu",
            font=("Segoe UI", 11, "bold"),
            foreground=Theme.TEXT_PRIMARY,
            background=Theme.BG_RESULTS_HEADER,
            anchor="w",
            padx=15
        )
        results_label.pack(side="left", fill="y")
        
        self.etl_loading_frame = tk.Frame(left_frame, bg="#ffffff", relief="solid", borderwidth=2, highlightbackground="#3498db", highlightthickness=1)
        
        etl_inner = tk.Frame(self.etl_loading_frame, bg=Theme.BG_WHITE, padx=15, pady=12)
        etl_inner.pack(fill="both", expand=True)
        
        # Label và progress bar
        self.etl_label = tk.Label(
            etl_inner,
            text="⏱️ Đang lấy dữ liệu từ web...",
            font=("Segoe UI", 10, "bold"),
            bg=Theme.BG_WHITE,
            fg=Theme.PRIMARY,
            anchor="w"
        )
        self.etl_label.pack(side="top", fill="x", pady=(0, 8))
        
        progress_row = tk.Frame(etl_inner, bg=Theme.BG_WHITE)
        progress_row.pack(side="top", fill="x", pady=(0, 5))
        
        self.etl_progress = ttk.Progressbar(
            progress_row,
            mode="indeterminate",
            length=400,
            style="ETL.TProgressbar"
        )
        self.etl_progress.pack(side="left", fill="x", expand=True)
        
        # Nút Hủy cho tra cứu hàng loạt – ẩn mặc định
        self.etl_cancel_button = tk.Button(
            progress_row,
            text="✕ Hủy",
            command=self._cancel_batch_search,
            bg=Theme.ERROR,
            fg="white",
            font=("Segoe UI", 9, "bold"),
            relief="flat",
            borderwidth=0,
            padx=10,
            pady=4,
            cursor="hand2",
            activebackground=Theme.ERROR_DARK,
            activeforeground="white"
        )
        self.etl_cancel_button.pack(side="left", padx=(8, 0))
        self.etl_cancel_button.pack_forget()  # ban đầu ẩn đi
        
        # Status text nhỏ
        self.etl_status_var = tk.StringVar(value="Đang kết nối...")
        self.etl_status = tk.Label(
            etl_inner,
            textvariable=self.etl_status_var,
            font=("Segoe UI", 9),
            bg=Theme.BG_WHITE,
            fg=Theme.TEXT_SECONDARY,
            anchor="w"
        )
        self.etl_status.pack(side="top", fill="x")

        # Bảng kết quả với scrollbar - có border và shadow đẹp
        self.tree_frame = tk.Frame(left_frame, bg=Theme.BG_WHITE, relief="solid", borderwidth=2, highlightbackground=Theme.BORDER_TREE, highlightthickness=1)
        self.tree_frame.pack(side="top", fill="both", expand=True, padx=5)
        
        # Empty state label - hiển thị khi chưa có dữ liệu
        self.empty_label = tk.Label(
            self.tree_frame,
            text=MSG_EMPTY_STATE,
            font=("Segoe UI", 11, "italic"),
            fg=Theme.TEXT_SECONDARY,
            bg=Theme.BG_WHITE,
            justify="center"
        )
        self.empty_label.place(relx=0.5, rely=0.5, anchor="center")

        columns = ("name", "tax_code", "representative", "tax_address", "phone", "status", "operation_date", "business_type", "main_business")
        self.tree = ttk.Treeview(
            self.tree_frame,
            columns=columns,
            show="headings",
            height=15
        )
        
        # Cấu hình cột với sort
        column_configs = [
            ("name", "Tên công ty", 200),
            ("tax_code", "Mã số thuế", 120),
            ("representative", "Người đại diện", 150),
            ("tax_address", "Địa chỉ Thuế", 250),
            ("phone", "Điện thoại", 120),
            ("status", "Tình trạng", 120),
            ("operation_date", "Ngày hoạt động", 120),
            ("business_type", "Loại hình DN", 180),
            ("main_business", "Ngành nghề chính", 200),
        ]
        
        for col, text, width in column_configs:
            self.tree.heading(col, text=text, command=lambda c=col: self.sort_by_column(c, False))
            self.tree.column(col, width=width)

        # Scrollbar
        scrollbar = ttk.Scrollbar(self.tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Bind mouse wheel để scroll treeview
        def _on_tree_mousewheel(event):
            if event.delta:
                delta = -1 * (event.delta / 120)  # Windows: 120 units per notch
                self.tree.yview_scroll(int(delta), "units")
        
        def _on_tree_mousewheel_linux_up(event):
            self.tree.yview_scroll(-1, "units")
        
        def _on_tree_mousewheel_linux_down(event):
            self.tree.yview_scroll(1, "units")
        
        # Bind mouse wheel cho treeview và tree_frame
        self.tree.bind("<MouseWheel>", _on_tree_mousewheel)
        self.tree_frame.bind("<MouseWheel>", _on_tree_mousewheel)
        
        # Linux support (Button-4 = scroll up, Button-5 = scroll down)
        self.tree.bind("<Button-4>", _on_tree_mousewheel_linux_up)
        self.tree.bind("<Button-5>", _on_tree_mousewheel_linux_down)
        self.tree_frame.bind("<Button-4>", _on_tree_mousewheel_linux_up)
        self.tree_frame.bind("<Button-5>", _on_tree_mousewheel_linux_down)
        
        # Bind selection event
        self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)

        # Zebra rows với màu đẹp hơn và hover effect
        self.tree.tag_configure("even", background=Theme.BG_TREE_EVEN)
        self.tree.tag_configure("odd", background=Theme.BG_TREE_ODD)
        self.tree.tag_configure("selected", background=Theme.SELECTED, foreground=Theme.SELECTED_TEXT)

        # Hint text với background đẹp hơn
        hint_frame = tk.Frame(left_frame, bg=Theme.BG_HINT, relief="flat", borderwidth=1, highlightbackground=Theme.BORDER_ETL)
        hint_frame.pack(side="top", fill="x", padx=5, pady=(10, 0))
        hint = tk.Label(
            hint_frame, 
            text="💡 Mẹo: Double-click vào dòng để mở chi tiết trên web. Click vào header để sắp xếp.",
            foreground="#2980b9",
            font=("Segoe UI", 9),
            background="#d5e8f7",
            anchor="w",
            padx=15,
            pady=10
        )
        hint.pack(side="left", fill="x", expand=True)

        # map id → detail_url và detail data
        self.item_url_map: Dict[str, str] = {}  # tree item_id -> detail_url
        self.item_data_map: Dict[str, CompanyDetailRecord] = {}  # tree item_id -> detail data

        # Bind events
        self.tree.bind("<Double-1>", self.on_row_double_click)
        self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)
    
    def _build_detail_frame(self) -> None:
        """Xây dựng frame thông tin chi tiết"""
        right_frame = ttk.LabelFrame(self.main_pane, text="📋 Thông tin chi tiết", padding=20)
        self.main_pane.add(right_frame, weight=2)
        
        # Background cho right frame
        right_frame.configure(style="TLabelframe")
        
        # Scrollable frame cho details
        detail_canvas = tk.Canvas(right_frame, bg="#ffffff", highlightthickness=0)
        detail_scrollbar = ttk.Scrollbar(right_frame, orient="vertical", command=detail_canvas.yview)
        detail_scrollable_frame = tk.Frame(detail_canvas, bg="#ffffff")
        
        detail_scrollable_frame.bind(
            "<Configure>",
            lambda e: detail_canvas.configure(scrollregion=detail_canvas.bbox("all"))
        )
        
        detail_canvas.create_window((0, 0), window=detail_scrollable_frame, anchor="nw")
        detail_canvas.configure(yscrollcommand=detail_scrollbar.set)
        
        # Lưu reference để có thể bind mouse wheel
        self.detail_canvas = detail_canvas
        self.detail_scrollable_frame = detail_scrollable_frame
        
        # Bind mouse wheel để scroll
        def _on_mousewheel(event):
            # Windows và MacOS: event.delta
            if event.delta:
                delta = -1 * (event.delta / 120)  # Windows: 120 units per notch
                detail_canvas.yview_scroll(int(delta), "units")
        
        def _on_mousewheel_linux_up(event):
            detail_canvas.yview_scroll(-1, "units")
        
        def _on_mousewheel_linux_down(event):
            detail_canvas.yview_scroll(1, "units")
        
        # Bind mouse wheel cho canvas và scrollable frame
        detail_canvas.bind("<MouseWheel>", _on_mousewheel)
        detail_scrollable_frame.bind("<MouseWheel>", _on_mousewheel)
        
        # Linux support (Button-4 = scroll up, Button-5 = scroll down)
        detail_canvas.bind("<Button-4>", _on_mousewheel_linux_up)
        detail_canvas.bind("<Button-5>", _on_mousewheel_linux_down)
        detail_scrollable_frame.bind("<Button-4>", _on_mousewheel_linux_up)
        detail_scrollable_frame.bind("<Button-5>", _on_mousewheel_linux_down)
        
        # Bind mouse wheel cho tất cả widgets con trong scrollable frame
        def bind_mousewheel_recursive(widget):
            widget.bind("<MouseWheel>", _on_mousewheel)
            widget.bind("<Button-4>", _on_mousewheel_linux_up)
            widget.bind("<Button-5>", _on_mousewheel_linux_down)
            for child in widget.winfo_children():
                bind_mousewheel_recursive(child)
        
        # Bind cho scrollable frame và tất cả children hiện tại
        bind_mousewheel_recursive(detail_scrollable_frame)
        
        detail_canvas.pack(side="left", fill="both", expand=True)
        detail_scrollbar.pack(side="right", fill="y")

        # Các label hiển thị thông tin chi tiết
        self.detail_labels = {}
        detail_fields = [
            ("name", "Tên công ty:"),
            ("tax_code", "Mã số thuế:"),
            ("representative", "Người đại diện:"),
            ("tax_address", "Địa chỉ Thuế:"),
            ("address", "Địa chỉ:"),
            ("phone", "Điện thoại:"),
            ("status", "Tình trạng:"),
            ("operation_date", "Ngày hoạt động:"),
            ("managed_by", "Quản lý bởi:"),
            ("business_type", "Loại hình DN:"),
            ("main_business", "Ngành nghề chính:"),
            ("other_businesses", "Ngành nghề khác:"),
        ]
        
        for field, label_text in detail_fields:
            # Frame với border nhẹ
            frame = tk.Frame(detail_scrollable_frame, bg=Theme.BG_WHITE, relief="flat")
            frame.pack(side="top", fill="x", pady=4, padx=5)
            
            # Separator line
            separator = tk.Frame(frame, height=1, bg=Theme.BG_RESULTS_HEADER)
            separator.pack(side="top", fill="x", pady=(0, 8))
            
            label = tk.Label(
                frame, 
                text=label_text, 
                font=("Segoe UI", 9, "bold"), 
                width=20, 
                anchor="w",
                bg=Theme.BG_WHITE,
                fg=Theme.TEXT_PRIMARY
            )
            label.pack(side="left", padx=(5, 10))
            
            # Tăng wraplength cho các field có thể dài như "Ngành nghề khác"
            wraplength = 350 if field == "other_businesses" else 280
            value_label = tk.Label(
                frame, 
                text="", 
                font=("Segoe UI", 9), 
                foreground="#34495e",  # Slightly darker than TEXT_PRIMARY for emphasis
                bg=Theme.BG_WHITE,
                anchor="w",
                wraplength=wraplength,
                justify="left"
            )
            value_label.pack(side="left", fill="x", expand=True, padx=(0, 5))
            
            # Bind mouse wheel cho frame và labels để scroll hoạt động khi hover vào
            def _on_detail_mousewheel(event):
                if event.delta:
                    detail_canvas.yview_scroll(-1 * int(event.delta / 120), "units")
            
            def _on_detail_mousewheel_linux_up(event):
                detail_canvas.yview_scroll(-1, "units")
            
            def _on_detail_mousewheel_linux_down(event):
                detail_canvas.yview_scroll(1, "units")
            
            frame.bind("<MouseWheel>", _on_detail_mousewheel)
            frame.bind("<Button-4>", _on_detail_mousewheel_linux_up)
            frame.bind("<Button-5>", _on_detail_mousewheel_linux_down)
            label.bind("<MouseWheel>", _on_detail_mousewheel)
            label.bind("<Button-4>", _on_detail_mousewheel_linux_up)
            label.bind("<Button-5>", _on_detail_mousewheel_linux_down)
            value_label.bind("<MouseWheel>", _on_detail_mousewheel)
            value_label.bind("<Button-4>", _on_detail_mousewheel_linux_up)
            value_label.bind("<Button-5>", _on_detail_mousewheel_linux_down)
            
            self.detail_labels[field] = value_label

    def set_status(self, text: str, force_update: bool = False) -> None:
        """
        Log status message (status bar đã bị xóa, dùng logging thay thế).
        
        Args:
            text: Text status để log
            force_update: Không dùng nữa (giữ lại để tương thích)
        """
        # Log status message thay vì hiển thị trên status bar (đã bị xóa)
        logger.info(f"[UI Status] {text}")
        logger.info(f"[UI Status] {text}")

    def _on_query_change(self, event: Optional[tk.Event] = None) -> None:
        """Xử lý khi người dùng thay đổi query - hiển thị validation"""
        
        query = self.query_var.get()
        if not query:
            self.query_status_label.config(text="", foreground="#3498db")
            return

        # Kiểm tra nếu là MST
        cleaned_query = ''.join(c for c in query if c.isdigit())
        if cleaned_query and is_valid_tax_code(cleaned_query):
            self.query_status_label.config(text=MSG_SEARCHING_MST, foreground=Theme.STATUS_SUCCESS)
        else:
            # Kiểm tra độ dài
            sanitized = sanitize_query(query)
            if len(query) > MAX_QUERY_LENGTH:
                self.query_status_label.config(
                    text=f"⚡ Query quá dài ({len(query)}/{MAX_QUERY_LENGTH} ký tự)",
                    foreground=Theme.STATUS_ERROR
                )
            else:
                self.query_status_label.config(text="", foreground=Theme.STATUS_INFO)

    def on_search(self) -> None:
        """Xử lý sự kiện tra cứu nhanh"""
        
        # Kiểm tra nếu đang batch thì không cho tra cứu nhanh
        if self._batch_state == "RUNNING":
            messagebox.showwarning(
                "Đang tra cứu hàng loạt",
                "Vui lòng đợi quá trình tra cứu hàng loạt hoàn thành hoặc hủy nó trước."
            )
            return
        
        query = self.query_var.get()
        original_query = query
        
        # Kiểm tra độ dài trước khi sanitize
        if len(query) > MAX_QUERY_LENGTH:
            if not messagebox.askyesno(
                "Query quá dài",
                f"Query của bạn có {len(query)} ký tự (tối đa {MAX_QUERY_LENGTH}).\n\n"
                f"Bạn có muốn cắt bớt và tiếp tục không?"
            ):
                return
            query = query[:MAX_QUERY_LENGTH]
        
        query = sanitize_query(query)
        
        if not query:
            messagebox.showwarning("Thiếu dữ liệu", ERR_INVALID_INPUT)
            return

        # Kiểm tra nếu là MST để hiển thị thông báo
        cleaned_query = ''.join(c for c in query if c.isdigit())
        is_tax_code = cleaned_query and is_valid_tax_code(cleaned_query)
        
        self.set_status("⏱️ Đang tra cứu..." + (" (theo MST)" if is_tax_code else ""))
        self.tree.delete(*self.tree.get_children())
        self.item_url_map.clear()
        self.item_data_map.clear()
        
        self.progress_queue.put(("show_loading", "Đang kết nối..."))

        threading.Thread(
            target=self._search_in_thread,
            args=(query,),
            daemon=True
        ).start()
    
    def _search_in_thread(self, query: str) -> None:
        """Tra cứu trong thread riêng - thread-safe với queue"""
        try:
            # Cập nhật status qua queue
            self.progress_queue.put(("etl_status", "Đang tìm kiếm thông tin công ty..."))
            results = self.client.search_companies(query=query, page=1)
            
            # Nếu có kết quả và cần fetch details
            if results and results[0].detail_url:
                self.progress_queue.put(("etl_status", "Đang lấy thông tin chi tiết từ web..."))
            
            # Cập nhật UI từ main thread
            # Capture giá trị ngay lúc tạo lambda để tránh closure issue
            self.after(0, lambda r=results, q=query: self._display_search_results(r, q))
        except CaptchaRequiredError as e:
            error = e
            self.after(0, lambda err=error: self._handle_captcha_error_single(err))
        except NetworkError as e:
            error = e
            self.after(0, lambda err=error: self._handle_network_error(err))
        except ValidationError as e:
            error = e
            self.after(0, lambda err=error: self._handle_validation_error(err))
        except (NetworkError, ParseError) as e:
            error = e
            self.after(0, lambda err=error: self._handle_search_error(err))
        except Exception as e:
            # Catch-all cho các lỗi không mong đợi (logic errors, etc.)
            # Log để debug nhưng vẫn hiển thị cho user
            logger.exception("Unexpected error in search thread")
            error = e
            self.after(0, lambda err=error: self._handle_search_error(err))
    
    def _display_search_results(self, results: List[CompanySearchResult], query: str) -> None:
        """Hiển thị kết quả tra cứu (chạy từ main thread)"""
        # Ẩn ETL loading
        self._hide_loading()

        # Xóa kết quả cũ
        self.tree.delete(*self.tree.get_children())
        self.item_url_map.clear()
        self.item_data_map.clear()

        # Không có kết quả → empty state
        if not results:
            if hasattr(self, "empty_label"):
                self.empty_label.config(
                    text=(
                        f"✗ Không tìm thấy kết quả cho: {query}\n\n"
                        "Hãy thử từ khóa khác hoặc kiểm tra lại mã số thuế."
                    )
                )
                self.empty_label.place(relx=0.5, rely=0.5, anchor="center")


            for label in self.detail_labels.values():
                label.config(text="", fg=Theme.TEXT_PRIMARY)

            self.set_status("✗ Không tìm thấy kết quả phù hợp.", force_update=True)
            return

        if hasattr(self, "empty_label"):
            self.empty_label.place_forget()

        for index, r in enumerate(results):
            tag = "even" if index % 2 == 0 else "odd"
            item_id = self.tree.insert(
                "",
                "end",
                values=(
                    r.name,
                    r.tax_code,
                    r.representative or "",
                    r.tax_address or r.address or "",
                    r.phone or "",
                    r.status or "",
                    r.operation_date or "",
                    r.business_type or "",
                    r.main_business or "",
                ),
                tags=(tag,),
            )

            if r.detail_url:
                self.item_url_map[item_id] = r.detail_url

            self.item_data_map[item_id] = {
                "name": r.name,
                "tax_code": r.tax_code,
                "representative": r.representative or "",
                "tax_address": r.tax_address or "",
                "address": r.address or "",
                "phone": r.phone or "",
                "status": r.status or "",
                "operation_date": r.operation_date or "",
                "managed_by": r.managed_by or "",
                "business_type": r.business_type or "",
                "main_business": r.main_business or "",
                "other_businesses": r.other_businesses or "",
            }

        # Cập nhật status

        self.set_status(
            MSG_SEARCH_SUCCESS.format(count=len(results)),
            force_update=True,
        )

        # Auto select dòng đầu để hiển thị chi tiết
        first = self.tree.get_children()
        if first:
            first_item = first[0]
            self.tree.selection_set(first_item)
            self.tree.focus(first_item)
            self._on_tree_select()
    
    def _handle_captcha_error_single(self, error: CaptchaRequiredError) -> None:
        """Xử lý lỗi CAPTCHA cho tra cứu đơn"""
        self._hide_loading()
        self.set_status("⚡ Website yêu cầu xác minh")
        messagebox.showwarning(
            "Website yêu cầu xác minh",
            "🔒 Server yêu cầu CAPTCHA\n\n"
            "Đây là hạn chế bảo mật của website để chống bot tự động.\n"
            "Ứng dụng chỉ hỗ trợ phát hiện CAPTCHA, không giải tự động.\n\n"
            "💡 Hướng dẫn:\n"
            "1. Mở trình duyệt và truy cập masothue.com\n"
            "2. Giải CAPTCHA thủ công trên website\n"
            "3. Đợi vài phút rồi thử lại trong ứng dụng\n"
            "4. Hoặc tra cứu ít dòng hơn để tránh bị chặn\n\n"
            "⚡ Lưu ý: Đừng tra quá dày để tránh bị chặn."
        )
    
    def _handle_network_error(self, error: NetworkError) -> None:
        """Xử lý lỗi network cho tra cứu đơn"""
        self._hide_loading()
        self.set_status("✗ Lỗi kết nối")
        messagebox.showerror(
            "Lỗi kết nối",
            f"Không thể kết nối đến server.\n\n"
            f"Chi tiết: {error.message}\n\n"
            f"Vui lòng kiểm tra kết nối internet và thử lại."
        )
    
    def _handle_validation_error(self, error: ValidationError) -> None:
        """Xử lý lỗi validation cho tra cứu đơn"""
        self._hide_loading()
        self.set_status(MSG_VALIDATION_ERROR)
        messagebox.showerror(
            "Dữ liệu không hợp lệ",
            f"{error.message}\n\n"
            f"Vui lòng kiểm tra lại thông tin nhập vào."
        )
    
    def _handle_search_error(self, error: Exception) -> None:
        """Xử lý lỗi tra cứu (chạy từ main thread)"""
        self.progress_queue.put(("hide_loading", None))
        self.set_status("✗ Lỗi")
        messagebox.showerror(TITLE_ERROR, MSG_SEARCH_ERROR_DETAIL.format(error=error))
    
    def _show_loading(self, status_text: str = "Đang kết nối...") -> None:
        """Hiển thị thanh ETL loading - rõ ràng khi đang lấy dữ liệu từ web"""
        # Luôn pack lại frame để đảm bảo hiển thị (nếu đã pack thì pack lại sẽ không lỗi)
        try:
            self.etl_loading_frame.pack_forget()  # Xóa pack cũ nếu có
        except:
            pass
        # Pack lại frame - đặt ngay sau results_header, trước tree_frame
        # Sử dụng pack(before=...) để không làm ảnh hưởng đến vị trí của results_header
        try:
            if hasattr(self, "tree_frame"):
                self.etl_loading_frame.pack(side="top", fill="x", padx=5, pady=(0, 10), before=self.tree_frame)
            else:
                self.etl_loading_frame.pack(side="top", fill="x", padx=5, pady=(0, 10))
        except (tk.TclError, AttributeError):
            self.etl_loading_frame.pack(side="top", fill="x", padx=5, pady=(0, 10))
        self.update_idletasks()  # Force update để frame hiển thị ngay
        
        self.etl_progress.start(15)  # Animation speed
        self.etl_status_var.set(status_text)
        self._custom_status_set = False
        # Animation cho label
        self._loading_spinner_index = 0
        self._animate_etl_loading()
    
    def _poll_progress_queue(self) -> None:
        """Poll progress queue từ main thread - thread-safe"""
        try:
            while True:
                msg_type, data = self.progress_queue.get_nowait()
                if msg_type == "progress":
                    # Progress được hiển thị trong card ETL loading, không cần xử lý gì thêm
                    pass
                elif msg_type == "etl_status":
                    status_text = data
                    self._update_etl_status(status_text)
                elif msg_type == "show_loading":
                    self._show_loading(data if data else "Đang kết nối...")
                elif msg_type == "hide_loading":
                    self._hide_loading()
        except queue.Empty:
            pass
        # Tiếp tục polling mỗi 100ms
        self.after(100, self._poll_progress_queue)
    
    def _update_etl_status(self, status_text: str) -> None:
        """Cập nhật status text cho ETL loading"""
        self.etl_status_var.set(status_text)
        self._custom_status_set = True
    
    def _hide_loading(self) -> None:
        """Ẩn thanh ETL loading"""
        self.etl_progress.stop()
        self.etl_loading_frame.pack_forget()
        
        # Ẩn luôn nút Hủy nếu đang hiển thị
        if hasattr(self, "etl_cancel_button"):
            self.etl_cancel_button.pack_forget()
        
        if hasattr(self, "_loading_animation_id"):
            self.after_cancel(self._loading_animation_id)
        self._custom_status_set = False
    
    def _animate_etl_loading(self) -> None:
        """Animation cho ETL loading label - thay đổi text động"""
        spinner_chars = ["⏱️", "🔄", "⏱️", "🔄"]
        status_texts = [
            "Đang kết nối đến server...",
            "Đang tải dữ liệu từ web...",
            "Đang xử lý thông tin...",
            "Đang phân tích kết quả..."
        ]
        if hasattr(self, '_loading_spinner_index'):
            char = spinner_chars[self._loading_spinner_index % len(spinner_chars)]
            status = status_texts[self._loading_spinner_index % len(status_texts)]
            self.etl_label.config(text=f"{char} Đang lấy dữ liệu từ web...")
            # Chỉ cập nhật status nếu chưa có status cụ thể từ thread
            if not hasattr(self, '_custom_status_set') or not self._custom_status_set:
                self.etl_status_var.set(status)
            self._loading_spinner_index += 1
            self._loading_animation_id = self.after(500, self._animate_etl_loading)

    def sort_by_column(self, col: str, descending: bool) -> None:
        """Sort cột khi click header"""
        data = [(self.tree.set(child, col), child) for child in self.tree.get_children("")]
        
        # Sắp xếp (xử lý số và text)
        try:
            data.sort(key=lambda x: (float(x[0]) if x[0] and x[0].replace('.', '').replace('-', '').isdigit() else float('inf'), x[0]), reverse=descending)
        except:
            data.sort(key=lambda x: x[0] or "", reverse=descending)
        
        # Di chuyển items
        for index, (val, child) in enumerate(data):
            self.tree.move(child, "", index)
            # Giữ nguyên tag (zebra rows)
            tags = self.tree.item(child, "tags")
            if tags:
                tag = "even" if index % 2 == 0 else "odd"
                self.tree.item(child, tags=(tag,))
        
        # Toggle ascending/descending
        self.tree.heading(col, command=lambda c=col: self.sort_by_column(c, not descending))
    
    def _on_tree_select(self, event: Optional[tk.Event] = None) -> None:
        """Xử lý khi chọn dòng trong tree"""
        selection = self.tree.selection()
        if not selection:
            # Không có selection: reset panel chi tiết
            for label in self.detail_labels.values():
                label.config(text="", fg=Theme.TEXT_PRIMARY)
            return
        # Có selection: gọi on_row_select
        self.on_row_select(event)
    
    def on_row_select(self, event: Optional[tk.Event] = None) -> None:
        """Hiển thị thông tin chi tiết khi chọn dòng"""
        item_id = self.tree.focus()
        if not item_id:
            for label in self.detail_labels.values():
                label.config(text="", fg=Theme.TEXT_PRIMARY)
            return
        
        data = self.item_data_map.get(item_id, {})
        
        for field, label in self.detail_labels.items():
            value = data.get(field, "")
            if value:
                label.config(text=value, fg="#2c3e50")
            else:
                label.config(text="(Chưa có thông tin)", fg="#95a5a6")

    def on_row_double_click(self, event: Optional[tk.Event] = None) -> None:
        """Mở web khi double-click"""
        item_id = self.tree.focus()
        if not item_id:
            return
        url = self.item_url_map.get(item_id)
        if url:
            webbrowser.open(url)
        else:
            messagebox.showinfo(
                "Thông tin",
                "Không có URL chi tiết cho bản ghi này. Bạn có thể copy MST và tìm trực tiếp trên web."
            )

    def on_import_excel(self) -> None:
        """Mở dialog chọn file Excel và tra cứu hàng loạt"""
        # Khóa UI nếu đang batch
        if self._batch_state == "RUNNING":
            messagebox.showwarning(
                "Đang tra cứu hàng loạt",
                "Vui lòng đợi quá trình tra cứu hàng loạt hiện tại hoàn thành hoặc hủy nó trước."
            )
            return
        
        # UI: Chọn file
        file_path = filedialog.askopenfilename(
            title="Chọn file Excel",
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("All files", "*.*")
            ],
            initialdir=self.last_dir
        )
        if not file_path:
            return
        
        # Lưu thư mục để lần sau
        self.last_dir = str(Path(file_path).parent)

        try:
            queries, query_column_idx, query_column = read_queries_from_excel(file_path)
            
            if query_column_idx is None:
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                headers = []
                if ws.max_row > 0:
                    for cell in ws[1]:
                        headers.append(cell.value if cell.value else "")
                
                    query_column_idx, query_column = self._show_column_selection_dialog(headers)
                    if query_column_idx is None:
                        return
                    
                    queries, _, query_column = read_queries_from_excel(file_path, query_column_idx=query_column_idx)

            if not queries:
                messagebox.showwarning(TITLE_WARNING, MSG_NO_EXCEL_DATA)
                return

            # UI: Xác nhận với người dùng
            confirm = messagebox.askyesno(
                TITLE_CONFIRM,
                CONFIRM_EXPORT.format(count=len(queries), column=query_column or "đã chọn")
            )
            if not confirm:
                return

            # Chuẩn bị trạng thái batch
            self.batch_results = []
            
            # UI: Chuyển sang state RUNNING
            self._set_batch_state(
                "RUNNING",
                status_text=f"Đang bắt đầu tra cứu {len(queries)} công ty..."
            )
            
            self.set_status(f"⏱️ Đang tra cứu {len(queries)} công ty...", force_update=True)
            
            # Chạy tra cứu hàng loạt trong thread riêng
            threading.Thread(
                target=self._batch_search,
                args=(queries, query_column or "đã chọn"),
                daemon=True
            ).start()

        except ValidationError as e:
            messagebox.showerror(TITLE_ERROR, MSG_FILE_INVALID.format(error=e.message))
        except FileError as e:
            messagebox.showerror(TITLE_ERROR, MSG_FILE_READ_ERROR.format(error=e.message))
        except Exception as e:
            logger.exception("Unexpected error in on_import_excel")
            messagebox.showerror("Lỗi đọc file Excel", ERR_FILE_READ.format(error=str(e)))
    
    def _show_column_selection_dialog(self, headers: List[str]) -> tuple[Optional[int], Optional[str]]:
        """
        Hiển thị dialog cho người dùng chọn cột để tra cứu.
        
        Args:
            headers: Danh sách tên cột từ Excel
            
        Returns:
            Tuple (column_index, column_name) hoặc (None, None) nếu người dùng hủy
        """
        if not headers:
            messagebox.showwarning(
                MSG_NO_COLUMNS,
                MSG_NO_COLUMNS_DETAIL
            )
            return None, None
        
        dialog = tk.Toplevel(self)
        dialog.title("Chọn cột để tra cứu")
        dialog.geometry("500x400")
        dialog.transient(self)
        dialog.grab_set()
        
        # Frame chính
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill="both", expand=True)
        
        info_label = tk.Label(
            main_frame,
            text=MSG_COLUMN_SELECTION_INFO,
            font=("Segoe UI", 10),
            justify="left",
            wraplength=450
        )
        info_label.pack(anchor="w", pady=(0, 15))
        
        list_frame = ttk.Frame(main_frame)
        list_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side="right", fill="y")
        
        listbox = tk.Listbox(
            list_frame,
            font=("Segoe UI", 10),
            height=10,
            yscrollcommand=scrollbar.set
        )
        listbox.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=listbox.yview)
        
        # Thêm headers vào listbox
        for idx, header in enumerate(headers):
            display_text = f"Cột {idx + 1}: {header}" if header else f"Cột {idx + 1}: (trống)"
            listbox.insert("end", display_text)
        
        # Chọn item đầu tiên
        if listbox.size() > 0:
            listbox.selection_set(0)
            listbox.activate(0)
        
        # Biến để lưu kết quả
        result = [None, None]
        
        def on_ok():
            selection = listbox.curselection()
            if selection:
                selected_idx = selection[0]
                result[0] = selected_idx
                result[1] = headers[selected_idx] if headers[selected_idx] else f"Cột {selected_idx + 1}"
                dialog.destroy()
            else:
                messagebox.showwarning(MSG_COLUMN_NOT_SELECTED, MSG_COLUMN_NOT_SELECTED_DETAIL)
        
        def on_cancel():
            dialog.destroy()
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x")
        
        ttk.Button(button_frame, text="Hủy", command=on_cancel).pack(side="right", padx=5)
        ttk.Button(button_frame, text="OK", command=on_ok).pack(side="right", padx=5)
        
        # Bind double-click để chọn nhanh
        listbox.bind("<Double-1>", lambda e: on_ok())
        
        # Focus vào dialog
        dialog.focus_set()
        dialog.wait_window()
        
        return result[0], result[1]

    def _cancel_batch_search(self) -> None:
        """Yêu cầu hủy batch search (gửi tín hiệu cho thread dừng lại)"""
        # Nếu không có batch nào đang chạy thì bỏ qua
        if self._batch_state != "RUNNING":
            return

        if not messagebox.askyesno(
            "Xác nhận hủy",
            "Bạn có chắc muốn hủy quá trình tra cứu hàng loạt?\n\n"
            "Kết quả đã tra cứu xong vẫn được giữ lại."
        ):
            return

        # Bật cờ cho thread biết cần dừng
        self._batch_cancelled = True
        
        self._set_batch_state("CANCELLED")

        # Disable nút Hủy trên card ETL
        if hasattr(self, "etl_cancel_button"):
            self.etl_cancel_button.config(state="disabled")

        # Thông báo rõ ràng cho người dùng về cooperative cancel
        self.set_status(
            "⏹ Đã yêu cầu hủy tra cứu hàng loạt. "
            "Hệ thống sẽ dừng sau khi xử lý xong công ty hiện tại "
            "(tối đa ~8-17 giây nếu đang trong request).",
            force_update=True,
        )
    
    def _batch_search(self, queries: List[str], query_column: str) -> None:
        """
        Tra cứu hàng loạt trong thread riêng.
        UI layer: Chỉ orchestrate worker và cập nhật UI, không có business logic.
        """
        # Tạo worker instance
        worker = BatchWorker(self.client)
        
        # UI callbacks
        def progress_callback(idx: int, total: int, query: str) -> None:
            """Cập nhật progress (UI logic)"""
            self._update_batch_progress(idx, total, query)
        
        def result_callback(result_dict: BatchResultDict) -> None:
            """Lưu kết quả (UI logic)"""
            self.batch_results.append(result_dict)
        
        def error_callback(query: str, exception: Exception) -> None:
            """Xử lý lỗi (UI logic)"""
            logger.error(f"Error processing query '{query}': {str(exception)}")
        
        def captcha_callback(idx: int, total: int, error_msg: str) -> None:
            """Xử lý CAPTCHA (UI logic)"""
            self._notify_captcha_error(idx, total, error_msg)
        
        try:
            results = worker.process_queries(
                queries,
                cancelled_callback=lambda: self._batch_cancelled,
                progress_callback=progress_callback,
                result_callback=result_callback,
                error_callback=error_callback,
                captcha_callback=captcha_callback
            )
            
            if not self._batch_cancelled and not worker.is_cancelled:
                self._notify_batch_complete(len(queries))
            elif self._batch_cancelled or worker.is_cancelled:
                pass
                
        except Exception as e:
            logger.exception("Unexpected error in _batch_search")
            if self.batch_results:
                self._notify_batch_complete(len(self.batch_results))
    
    def _update_batch_progress(self, idx: int, total: int, query: str) -> None:
        """Cập nhật progress cho batch search (UI logic)"""
        self.progress_queue.put(("etl_status", f"Đang tra cứu: {query} ({idx}/{total})"))

    def _notify_batch_cancelled(self, completed: int, total: int) -> None:
        """Thông báo batch đã bị hủy (UI logic)"""
        self.progress_queue.put(("hide_loading", None))
        # Capture giá trị ngay lúc tạo lambda để tránh closure issue
        self.after(0, lambda completed=completed: self._batch_search_cancelled(completed, total))
    
    def _notify_captcha_error(self, idx: int, total: int, error_msg: str) -> None:
        """Thông báo lỗi CAPTCHA (UI logic)"""
        self.progress_queue.put(("hide_loading", None))
        # Capture giá trị ngay lúc tạo lambda để tránh closure issue
        self.after(0, lambda idx=idx, msg=error_msg: self._handle_captcha_error(idx, total, msg))
    
    def _notify_batch_complete(self, total: int) -> None:
        """Thông báo batch hoàn thành (UI logic)"""
        self.progress_queue.put(("hide_loading", None))
        # Capture giá trị ngay lúc tạo lambda để tránh closure issue
        self.after(0, lambda t=total: self._batch_search_complete(t))
    
    
    def _batch_search_cancelled(self, completed: int, total: int) -> None:
        """Xử lý khi batch search bị hủy (được gọi từ thread khi thực sự dừng)"""
        self._set_batch_state("CANCELLED")
        
        # Đảm bảo bảng hiển thị đúng số kết quả đã xử lý xong
        # (self.batch_results hiện đang chứa đúng completed phần tử)
        self._render_batch_results()
        
        # Hiện thông báo (có thể trễ vài giây sau khi bấm Hủy)
        messagebox.showinfo(
            "Đã hủy",
            SUCCESS_BATCH_CANCELLED.format(completed=completed, total=total)
        )
        
    
    def _handle_captcha_error(self, completed: int, total: int, error_msg: str) -> None:
        """Xử lý khi phát hiện CAPTCHA trong quá trình tra cứu"""
        self._batch_cancelled = True
        self._set_batch_state("CANCELLED")
        
        # Disable nút Hủy trên card ETL
        try:
            if hasattr(self, "etl_cancel_button"):
                self.etl_cancel_button.config(state="disabled")
        except (tk.TclError, AttributeError):
            pass  # Widget có thể đã bị destroy hoặc không tồn tại
        
        messagebox.showwarning(
            "Website yêu cầu xác minh",
            ERR_CAPTCHA.format(completed=completed, total=total)
        )
    
    def _set_batch_state(self, state: str, *, status_text: str = "") -> None:
        """
        State machine cho batch search - orchestrate state change và UI update.
        
        States:
            - "IDLE": Không có batch nào, UI mở khóa
            - "RUNNING": Đang chạy batch, UI khóa, hiển thị loading + nút Hủy
            - "CANCELLED": Đã hủy, UI mở khóa, ẩn loading
            - "FINISHED": Hoàn thành, UI mở khóa, ẩn loading
        
        Args:
            state: Một trong ["IDLE", "RUNNING", "CANCELLED", "FINISHED"]
            status_text: Text hiển thị trong ETL loading (dùng cho RUNNING)
        """
        self._batch_state = state
        self._update_batch_flags(state)
        self._update_ui_for_batch_state(state, status_text=status_text)
    
    def _update_batch_flags(self, state: str) -> None:
        """Cập nhật các flags nội bộ theo state (tách riêng logic state change)."""
        if state == "IDLE":
            self._is_batch_running = False
            self._batch_cancelled = False
        elif state == "RUNNING":
            self._is_batch_running = True
            self._batch_cancelled = False
        elif state == "CANCELLED":
            self._is_batch_running = False
            # _batch_cancelled giữ nguyên True (đã set trong _cancel_batch_search)
        elif state == "FINISHED":
            self._is_batch_running = False
            self._batch_cancelled = False
    
    def _update_ui_for_batch_state(self, state: str, *, status_text: str = "") -> None:
        """Cập nhật UI theo state (tách riêng logic UI update)."""
        if state == "IDLE":
            self._hide_loading_safe()
            self._unlock_ui_safe()
            self._hide_cancel_buttons_safe()
        
        elif state == "RUNNING":
            self._lock_ui_safe()
            self._show_loading_safe(status_text)
            self._show_cancel_button_safe()
        
        elif state == "CANCELLED":
            self._hide_loading_safe()
            self._unlock_ui_safe()
            self._hide_cancel_buttons_safe()
        
        elif state == "FINISHED":
            self._hide_loading_safe()
            self._unlock_ui_safe()
            self._hide_cancel_buttons_safe()
    
    def _lock_ui_safe(self) -> None:
        """Khóa các nút UI (safe với exception handling)."""
        try:
            self.search_button.config(state="disabled")
            self.query_entry.config(state="disabled")
            self.import_button.config(state="disabled")
            self.export_button.config(state="disabled")
        except (tk.TclError, AttributeError):
            pass
    
    def _unlock_ui_safe(self) -> None:
        """Mở khóa các nút UI (safe với exception handling)."""
        try:
            self.search_button.config(state="normal")
            self.query_entry.config(state="normal")
            self.import_button.config(state="normal")
            self.export_button.config(state="normal")
        except (tk.TclError, AttributeError):
            pass
    
    def _show_loading_safe(self, status_text: str = "") -> None:
        """Hiển thị ETL loading frame (safe với exception handling)."""
        try:
            self.etl_loading_frame.pack_forget()  # Xóa pack cũ nếu có
            # Pack lại frame - đặt ngay sau results_header, trước tree_frame
            try:
                if hasattr(self, "tree_frame"):
                    self.etl_loading_frame.pack(side="top", fill="x", padx=5, pady=(0, 10), before=self.tree_frame)
                else:
                    self.etl_loading_frame.pack(side="top", fill="x", padx=5, pady=(0, 10))
            except (tk.TclError, AttributeError):
                # Fallback nếu tree_frame không tồn tại
                self.etl_loading_frame.pack(side="top", fill="x", padx=5, pady=(0, 10))
            self.etl_progress.start(15)
            if status_text:
                self.etl_status_var.set(status_text)
            self.update_idletasks()
        except Exception:
            pass
    
    def _hide_loading_safe(self) -> None:
        """Ẩn ETL loading frame (safe với exception handling)."""
        try:
            self._hide_loading()
        except Exception:
            pass
    
    def _show_cancel_button_safe(self) -> None:
        """Hiển thị nút Hủy (safe với exception handling)."""
        try:
            if hasattr(self, "etl_cancel_button"):
                self.etl_cancel_button.config(state="normal")
                self.etl_cancel_button.pack(side="left", padx=(8, 0))
        except (tk.TclError, AttributeError):
            pass
    
    def _hide_cancel_buttons_safe(self) -> None:
        """Ẩn các nút Hủy (safe với exception handling)."""
        try:
            if hasattr(self, "etl_cancel_button"):
                self.etl_cancel_button.pack_forget()
            if hasattr(self, "cancel_button"):
                self.cancel_button.config(state="disabled")
            if hasattr(self, "close_button"):
                self.close_button.config(state="disabled")
        except (tk.TclError, AttributeError):
            pass
            
    
    def _render_batch_results(self) -> None:
        """Đổ self.batch_results vào tree view (dùng cho FINISHED / CANCELLED)."""
        # Xóa dữ liệu cũ
        self.tree.delete(*self.tree.get_children())
        self.item_url_map.clear()
        self.item_data_map.clear()
        
        # Hiển thị / ẩn empty state
        if hasattr(self, "empty_label"):
            if self.batch_results:
                self.empty_label.place_forget()
            else:
                self.empty_label.place(relx=0.5, rely=0.5, anchor="center")
        
        for index, r in enumerate(self.batch_results):
            tag = "even" if index % 2 == 0 else "odd"
            item_id = self.tree.insert(
                "",
                "end",
                values=(
                    r["Tên công ty"],
                    r["Mã số thuế"],
                    r.get("Người đại diện", ""),
                    r.get("Địa chỉ Thuế", r.get("Địa chỉ", "")),
                    r.get("Điện thoại", ""),
                    r.get("Tình trạng", ""),
                    r.get("Ngày hoạt động", ""),
                    r.get("Loại hình DN", ""),
                    r.get("Ngành nghề chính", "")
                ),
                tags=(tag,),
            )

            if r.get("URL"):
                self.item_url_map[item_id] = r["URL"]
            
            self.item_data_map[item_id] = {
                "name": r["Tên công ty"],
                "tax_code": r["Mã số thuế"],
                "representative": r.get("Người đại diện", ""),
                "tax_address": r.get("Địa chỉ Thuế", ""),
                "address": r.get("Địa chỉ", ""),
                "phone": r.get("Điện thoại", ""),
                "status": r.get("Tình trạng", ""),
                "operation_date": r.get("Ngày hoạt động", ""),
                "managed_by": r.get("Quản lý bởi", ""),
                "business_type": r.get("Loại hình DN", ""),
                "main_business": r.get("Ngành nghề chính", ""),
                "other_businesses": r.get("Ngành nghề khác", ""),
            }

        # Auto select dòng đầu để panel chi tiết cập nhật luôn
        children = self.tree.get_children()
        if children:
            first = children[0]
            self.tree.selection_set(first)
            self.tree.focus(first)
            self._on_tree_select()
    
    def _batch_search_complete(self, total: int) -> None:
        """Hoàn thành tra cứu hàng loạt (chạy từ main thread)"""
        self._set_batch_state("FINISHED")

        self.set_status(f"✓ Hoàn thành tra cứu {total} công ty.")

        # Đổ kết quả vào bảng (tất cả self.batch_results)
        self._render_batch_results()
        
        messagebox.showinfo(
            "Hoàn thành",
            f"Đã tra cứu xong {total} công ty.\n"
            f"Tìm thấy: {sum(1 for r in self.batch_results if r['Mã số thuế'])} kết quả.\n\n"
            f"Nhấn 'Xuất Excel' để lưu kết quả."
        )

    def on_export_excel(self) -> None:
        """Xuất kết quả tra cứu ra file Excel"""
        
        # UI: Thu thập dữ liệu từ UI
        # Ưu tiên dùng batch_results nếu có (đầy đủ hơn)
        query_source = "batch"
        if self.batch_results:
            export_data = self.batch_results
            query_source = "batch"
        else:
            # Nếu không có batch results, xuất từ tree view
            items = self.tree.get_children()
            if not items:
                messagebox.showwarning("Cảnh báo", "Không có dữ liệu để xuất.")
                return
            
            # Lấy từ tree view
            export_data = []
            query_source = "quick_search"
            for item_id in items:
                values = self.tree.item(item_id, "values")
                export_data.append({
                    "Tên công ty": values[0] if len(values) > 0 else "",
                    "Mã số thuế": values[1] if len(values) > 1 else "",
                    "Người đại diện": values[2] if len(values) > 2 else "",
                    "Địa chỉ Thuế": values[3] if len(values) > 3 else "",
                    "Điện thoại": values[4] if len(values) > 4 else "",
                    "Tình trạng": values[5] if len(values) > 5 else "",
                    "Ngày hoạt động": values[6] if len(values) > 6 else "",
                    "Loại hình DN": values[7] if len(values) > 7 else "",
                    "Ngành nghề chính": values[8] if len(values) > 8 else "",
                    "URL": self.item_url_map.get(item_id, "")
                })

        # UI: Chọn nơi lưu file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename_base = f"ket_qua_tra_cuu_{timestamp}"
        default_filename = sanitize_filename(default_filename_base) + ".xlsx"
        
        file_path = filedialog.asksaveasfilename(
            title="Lưu kết quả ra Excel",
            defaultextension=".xlsx",
            filetypes=[
                ("Excel files", "*.xlsx"),
                ("All files", "*.*")
            ],
            initialfile=default_filename,
            initialdir=self.last_dir
        )

        if not file_path:
            return
        
        # Sanitize filename (chỉ phần tên file, giữ nguyên đường dẫn)
        file_path_obj = Path(file_path)
        sanitized_name = sanitize_filename(file_path_obj.stem) + file_path_obj.suffix
        file_path = str(file_path_obj.parent / sanitized_name)
        
        # Lưu thư mục để lần sau
        self.last_dir = str(Path(file_path).parent)

        try:
            # Service layer: Xuất ra Excel
            metadata = {
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "source": "Tra cứu hàng loạt" if query_source == "batch" else "Tra cứu nhanh",
                "count": len(export_data)
            }
            
            export_results_to_excel(export_data, file_path, metadata)
            
            # UI: Thông báo thành công
            messagebox.showinfo("Thành công", f"Đã xuất kết quả ra file:\n{file_path}")
        except FileError as e:
            messagebox.showerror("Lỗi", f"Không xuất được file Excel.\nChi tiết: {e.message}")
        except Exception as e:
            logger.exception("Unexpected error in on_export_excel")
            messagebox.showerror("Lỗi", f"Không xuất được file Excel.\nChi tiết: {e}")
    
    def _on_closing(self) -> None:
        """Xử lý khi người dùng đóng cửa sổ"""
        # Kiểm tra xem có đang tra cứu hàng loạt không
        if self._batch_state == "RUNNING":
            # Đang tra cứu, hỏi xác nhận
            try:
                # Lấy thông tin từ etl_status_var hoặc batch_results
                if hasattr(self, "etl_status_var"):
                    status_text = self.etl_status_var.get()
                    match = re.search(r'\((\d+)/(\d+)\)', status_text)
                    if match:
                        completed = match.group(1)
                        total = match.group(2)
                        response = messagebox.askyesno(
                            "Xác nhận thoát",
                            f"Đang tra cứu hàng loạt ({completed}/{total}).\n\n"
                            f"Bạn có muốn dừng và thoát không?"
                        )
                    else:
                        completed = len(self.batch_results) if hasattr(self, "batch_results") else "?"
                        response = messagebox.askyesno(
                            "Xác nhận thoát",
                            f"Đang tra cứu hàng loạt.\n\n"
                            f"Bạn có muốn dừng và thoát không?"
                        )
                else:
                    response = messagebox.askyesno(
                        "Xác nhận thoát",
                        f"Đang tra cứu hàng loạt.\n\n"
                        f"Bạn có muốn dừng và thoát không?"
                    )
                
                if response:
                    # Đồng ý - dừng batch và thoát
                    self._batch_cancelled = True
                    # Thread đã được set daemon=True nên sẽ tự dừng khi app đóng
                    # Nhưng vẫn đợi một chút để thread có thể check cờ và dừng gracefully
                    def safe_destroy():
                        try:
                            self.destroy()
                        except Exception:
                            pass  # Ignore errors nếu widget đã bị destroy
                    self.after(1000, safe_destroy)  # Tăng thời gian đợi lên 1s
                    return
                else:
                    # Không đồng ý - không thoát
                    return
            except Exception:
                pass
        
        # Không đang batch hoặc đã xác nhận - thoát bình thường
        self.destroy()


if __name__ == "__main__":
    # Cấu hình logging ở entry point
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )
    
    app = MasothueApp()
    app.mainloop()
